#!/usr/bin/env python3
"""
Script to generate publications.xlsx with multiple sheets (Publications, Preprints, Talks).
This creates an Excel file that can be used by the website to load all content dynamically.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ===== PUBLICATIONS DATA =====
publications = [
    {
        "Year": "2025",
        "Venue": "Eurocrypt",
        "Title": "Pseudorandom Functions with Weak Programming Privacy and Applications to Private Information Retrieval",
        "Link": "https://eprint.iacr.org/2025/300",
        "Authors": "Ashrujit Ghoshal, Mingxun Zhou, Elaine Shi, Bo Peng",
        "Note": "(Randomized Author Order)",
        "CodeLink": ""
    },
    {
        "Year": "2025",
        "Venue": "ICLR",
        "Title": "Pacmann: Efficient Private Approximate Nearest Neighbor Search",
        "Link": "https://eprint.iacr.org/2024/1600",
        "Authors": "Mingxun Zhou, Elaine Shi, and Giulia Fanti",
        "Note": "",
        "CodeLink": ""
    },
    {
        "Year": "2024",
        "Venue": "CCS",
        "Title": "Conan: Distributed Proofs of Compliance for Anonymous Data Collection",
        "Link": "https://eprint.iacr.org/2023/1900",
        "Authors": "Mingxun Zhou, Elaine Shi, and Giulia Fanti",
        "Note": "",
        "CodeLink": ""
    },
    {
        "Year": "2024",
        "Venue": "Eurocrypt",
        "Title": "Efficient Pre-processing PIR Without Public-Key Cryptography",
        "Link": "https://eprint.iacr.org/2023/1574",
        "Authors": "Ashrujit Ghoshal, Mingxun Zhou, and Elaine Shi",
        "Note": "(Randomized Author Order)",
        "CodeLink": ""
    },
    {
        "Year": "2024",
        "Venue": "ITCS",
        "Title": "Advanced Composition Theorems for Differential Obliviousness",
        "Link": "https://eprint.iacr.org/2023/842",
        "Authors": "Mingxun Zhou, Mengshi Zhao, T-H. Hubert Chan, and Elaine Shi",
        "Note": "(Randomized Author Order)",
        "CodeLink": ""
    },
    {
        "Year": "2024",
        "Venue": "S&P",
        "Title": "Piano: Extremely Simple, Single-Server PIR with Sublinear Server Computation",
        "Link": "https://eprint.iacr.org/2023/452",
        "Authors": "Mingxun Zhou, Andrew Park, Elaine Shi and Wenting Zheng",
        "Note": "",
        "CodeLink": ""
    },
    {
        "Year": "2023",
        "Venue": "Eurocrypt",
        "Title": "A Theory of Composition for Differential Obliviousness",
        "Link": "https://eprint.iacr.org/2022/1357",
        "Authors": "Mingxun Zhou, Elaine Shi, T-H. Hubert Chan, and Shir Maimon",
        "Note": "(Randomized Author Order)",
        "CodeLink": ""
    },
    {
        "Year": "2023",
        "Venue": "Eurocrypt",
        "Title": "Optimal Single-Server Private Information Retrieval",
        "Link": "https://eprint.iacr.org/2022/609",
        "Authors": "Mingxun Zhou, W. Lin, Yiannis Tselekounis, and Elaine Shi",
        "Note": "(Randomized Author Order)",
        "CodeLink": ""
    },
    {
        "Year": "2023",
        "Venue": "INFOCOM",
        "Title": "Mercury: Fast Transaction Broadcast in High Performance Blockchain System",
        "Link": "https://wuwuz.github.io",
        "Authors": "Mingxun Zhou*, Liyi Zeng*, Yilin Han, Peilun Li, Fan Long, Dong Zhou, Ivan Beschastnikh, and Ming Wu",
        "Note": "(*Equal Contribution)",
        "CodeLink": ""
    },
    {
        "Year": "2022",
        "Venue": "S&P",
        "Title": "Locally Differentially Private Sparse Vector Aggregation",
        "Link": "https://arxiv.org/pdf/2112.03449.pdf",
        "Authors": "Mingxun Zhou, Tianhao Wang, T-H. Hubert Chan, Giulia Fanti, and Elaine Shi",
        "Note": "",
        "CodeLink": ""
    },
    {
        "Year": "2021",
        "Venue": "NDSS",
        "Title": "SquirRL: Automating Attack Analysis on Blockchain Incentive Mechanisms with Deep Reinforcement Learning",
        "Link": "https://arxiv.org/pdf/1912.01798.pdf",
        "Authors": "Charlie Hou*, Mingxun Zhou*, Yan Ji, Phil Daian, Florian Tramer, Giulia Fanti, and Ari Juels",
        "Note": "(*Equal Contribution)",
        "CodeLink": "https://github.com/wuwuz/SquirRL"
    },
    {
        "Year": "2020",
        "Venue": "VLDB",
        "Title": "Vacuum Filters: More Space-Efficient and Faster Replacement for Bloom and Cuckoo Filters",
        "Link": "http://www.vldb.org/pvldb/vol13/p197-wang.pdf",
        "Authors": "Minmei Wang*, Mingxun Zhou*, Shouqian Shi, and Chen Qian",
        "Note": "(*Equal Contribution)",
        "CodeLink": "https://github.com/wuwuz/Vacuum-Filter"
    }
]

# ===== PREPRINTS DATA =====
preprints = [
    {
        "Title": "PMark: Towards Robust and Distortion-free Semantic-level Watermarking with Channel Constraints",
        "Link": "https://arxiv.org/abs/2509.21057",
        "Authors": "Jiahao Huo, Shuliang Liu, Bin Wang, Junyan Zhang, Yibo Yan, Aiwei Liu, Xuming Hu, Mingxun Zhou",
        "Year": "2025",
        "Note": "",
        "Type": ""
    },
    {
        "Title": "Private Information Retrieval and Searching with Sublinear Costs",
        "Link": "https://csd.cmu.edu/sites/default/files/phd-thesis/CMU-CS-25-115.pdf",
        "Authors": "",
        "Year": "2025",
        "Note": "PhD Thesis",
        "Type": ""
    },
    {
        "Title": "The Power of the Differentially Oblivious Shuffle in Distributed Privacy Mechanisms",
        "Link": "https://eprint.iacr.org/2022/177.pdf",
        "Authors": "Mingxun Zhou, and Elaine Shi",
        "Year": "2022",
        "Note": "",
        "Type": ""
    },
    {
        "Title": "VRecon: An Efficient Set Reconciliation Algorithm",
        "Link": "",
        "Authors": "",
        "Year": "2021",
        "Note": "Bachelor Thesis",
        "Type": ""
    }
]

# ===== TALKS DATA =====
talks = [
    {
        "Title": "Advanced Composition Theorems for Differential Obliviousness",
        "Venue": "ITCS",
        "Date": "Jan. 2024",
        "VideoLink": "https://www.youtube.com/watch?v=oq1jIpUAy-0"
    },
    {
        "Title": "Recent Progress in Private Information Retrieval",
        "Venue": "PKU",
        "Date": "Jul. 2023",
        "VideoLink": "https://www.bilibili.com/video/BV1gF411X7BF"
    },
    {
        "Title": "Piano: Extremely Simple, Single-Server PIR with Sublinear Server Computation",
        "Venue": "CMU Crypto Seminar",
        "Date": "May 2023",
        "VideoLink": "https://www.youtube.com/watch?v=WOvk7grf2H0"
    },
    {
        "Title": "A Theory of Composition for Differential Obliviousness",
        "Venue": "Eurocrypt",
        "Date": "Apr 2023",
        "VideoLink": "https://www.youtube.com/watch?v=Jw11hbyFTCk"
    },
    {
        "Title": "Optimal Single-Server Private Information Retrieval",
        "Venue": "CMU Theory Lunch",
        "Date": "Sep. 2022",
        "VideoLink": "https://www.youtube.com/watch?v=E0t_iogzOR4"
    },
    {
        "Title": "The Power of the Differentially Oblivious Shuffle in Distributed Privacy Mechanisms",
        "Venue": "Symposium on Foundations of Responsible Computing",
        "Date": "Jun. 2022",
        "VideoLink": "https://www.youtube.com/watch?v=4WPkiz6yhvQ"
    },
    {
        "Title": "Locally Differentially Private Sparse Vector Aggregation",
        "Venue": "IEEE S&P",
        "Date": "May 2022",
        "VideoLink": "https://www.youtube.com/watch?v=d4TFqe7c_gs"
    },
    {
        "Title": "SquirRL: Automating Attack Analysis on Blockchain Incentive Mechanisms with Deep Reinforcement Learning",
        "Venue": "IJTCS",
        "Date": "Aug. 2020",
        "VideoLink": "https://www.bilibili.com/video/av202040234/"
    }
]

def create_sheet(wb, sheet_name, headers, data, column_widths):
    """Create a worksheet with headers and data."""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)
    
    # Write headers
    ws.append(headers)
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row in data:
        values = [row.get(h, "") for h in headers]
        ws.append(values)
    
    # Set column widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    return ws

def create_excel_file():
    """Create an Excel file with multiple sheets for Publications, Preprints, and Talks."""
    wb = Workbook()
    
    # Remove default sheet if we're creating multiple sheets
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create Publications sheet
    pub_headers = ["Year", "Venue", "Title", "Link", "Authors", "Note", "CodeLink"]
    pub_widths = {
        "A": 8,   # Year
        "B": 15,  # Venue
        "C": 80,  # Title
        "D": 50,  # Link
        "E": 60,  # Authors
        "F": 25,  # Note
        "G": 40   # CodeLink
    }
    create_sheet(wb, "Publications", pub_headers, publications, pub_widths)
    
    # Create Preprints sheet
    preprint_headers = ["Title", "Link", "Authors", "Year", "Note", "Type"]
    preprint_widths = {
        "A": 70,  # Title
        "B": 50,  # Link
        "C": 60,  # Authors
        "D": 8,   # Year
        "E": 20,  # Note
        "F": 20   # Type
    }
    create_sheet(wb, "Preprints", preprint_headers, preprints, preprint_widths)
    
    # Create Talks sheet
    talk_headers = ["Title", "Venue", "Date", "VideoLink"]
    talk_widths = {
        "A": 70,  # Title
        "B": 40,  # Venue
        "C": 15,  # Date
        "D": 50   # VideoLink
    }
    create_sheet(wb, "Talks", talk_headers, talks, talk_widths)
    
    # Save the file
    wb.save("publications.xlsx")
    print(f"✅ Created publications.xlsx with:")
    print(f"   - {len(publications)} publications")
    print(f"   - {len(preprints)} preprints")
    print(f"   - {len(talks)} talks")
    print("\nSheet structure:")
    print("\n📄 Publications sheet:")
    print("   Columns: Year, Venue, Title, Link, Authors, Note, CodeLink")
    print("\n📄 Preprints sheet:")
    print("   Columns: Title, Link, Authors, Year, Note, Type")
    print("\n📄 Talks sheet:")
    print("   Columns: Title, Venue, Date, VideoLink")

if __name__ == "__main__":
    try:
        create_excel_file()
    except ImportError:
        print("❌ Error: openpyxl library not found.")
        print("   Please install it with: pip install openpyxl")
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")

